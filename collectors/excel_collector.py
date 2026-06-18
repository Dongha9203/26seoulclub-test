"""
Excel(.xlsx) 파일 수집 모듈.
openpyxl을 사용해 행 단위로 Document를 생성합니다.
"""

import logging
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook

from models.document import Document
from utils.category_tagger import tag_category

logger = logging.getLogger(__name__)


def collect_excel(file_path: str, source_origin: Optional[str] = None) -> List[Document]:
    """
    .xlsx 파일을 파싱해 Document 리스트를 반환합니다.
    첫 번째 행을 헤더로 사용하고, 이후 각 행을 하나의 Document로 변환합니다.

    Args:
        file_path: .xlsx 파일 경로
        source_origin: 출처 이름 (미지정 시 파일명 사용)

    Raises:
        FileNotFoundError: 파일이 없을 경우
        ValueError: .xlsx가 아닌 파일일 경우
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")
    if path.suffix.lower() not in {".xlsx", ".xls"}:
        raise ValueError(
            f"지원하지 않는 파일 형식입니다: {path.suffix}\n"
            ".xlsx(Excel) 파일만 지원합니다."
        )

    origin = source_origin or path.name
    logger.info("Excel 수집 시작: %s", origin)

    wb = load_workbook(str(path), read_only=True, data_only=True)
    documents = []

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        # 첫 행을 헤더로 사용
        headers = [str(c).strip() if c is not None else f"컬럼{i}" for i, c in enumerate(rows[0])]

        # 제목/본문 컬럼 탐지
        title_col_idx = _find_col_idx(headers, ["질문", "title", "제목", "question", "항목"])
        content_col_idx = _find_col_idx(headers, ["답변", "content", "내용", "answer", "설명", "본문"])

        for row_idx, row in enumerate(rows[1:], start=2):
            cell_values = [str(c).strip() if c is not None else "" for c in row]
            if not any(cell_values):
                continue

            if title_col_idx is not None and content_col_idx is not None:
                title = cell_values[title_col_idx] if title_col_idx < len(cell_values) else ""
                content = cell_values[content_col_idx] if content_col_idx < len(cell_values) else ""
            elif headers:
                title = cell_values[0] if cell_values else ""
                content = " | ".join(
                    f"{headers[i]}: {v}"
                    for i, v in enumerate(cell_values[1:], start=1)
                    if v and i < len(headers)
                )
            else:
                continue

            if not title and not content:
                continue

            sheet_origin = f"{origin}:{sheet.title}" if len(wb.worksheets) > 1 else origin
            doc = Document.new(
                source_type="excel",
                source_origin=sheet_origin,
                title=title or f"{sheet.title} 행{row_idx}",
                content=content,
                is_editable=True,
            )
            doc.category = tag_category(doc.title, doc.content)
            documents.append(doc)

    wb.close()
    logger.info("Excel 수집 완료: %d개 Document 생성 (%s)", len(documents), origin)
    return documents


def _find_col_idx(headers: List[str], candidates: List[str]) -> Optional[int]:
    lower_headers = [h.lower() for h in headers]
    for c in candidates:
        if c.lower() in lower_headers:
            return lower_headers.index(c.lower())
    return None
